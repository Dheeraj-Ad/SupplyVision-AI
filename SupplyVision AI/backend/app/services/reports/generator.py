import io
import csv
from datetime import datetime, timezone, timedelta
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ReportLab Imports
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

from app.core.database import User, Organisation, AlertEvent, RecoveryPlan, AuditLog
from app.services.graph import graph_service
from app.services.signals import signals_service

# NumberedCanvas for professional "Page X of Y" footers
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#6B7280"))
        
        # Header (Only on page 2 and later)
        if self._pageNumber > 1:
            self.drawString(54, 750, "SupplyVision AI - Operations Resilience Digest")
            self.setStrokeColor(colors.HexColor("#E5E7EB"))
            self.setLineWidth(0.5)
            self.line(54, 742, 558, 742)
            
        # Footer
        footer_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(558, 40, footer_text)
        self.drawString(54, 40, "CONFIDENTIAL - For Internal Board Review Only")
        self.setStrokeColor(colors.HexColor("#E5E7EB"))
        self.setLineWidth(0.5)
        self.line(54, 52, 558, 52)
        
        self.restoreState()


class ReportGenerator:
    @staticmethod
    def _get_report_data(db: Session, org_id: str, report_type: str) -> dict:
        """Fetch matching dataset based on report type."""
        org = db.query(Organisation).filter(Organisation.id == org_id).first()
        org_name = org.name if org else "Supply Chain SME"
        
        # Fetch data points
        suppliers = graph_service.get_graph_data(org_id).get("nodes", [])
        suppliers = [n for n in suppliers if n.get("label") == "Supplier"]
        
        alerts = db.query(AlertEvent).filter(AlertEvent.org_id == org_id).all()
        plans = db.query(RecoveryPlan).filter(RecoveryPlan.org_id == org_id).all()
        audits = db.query(AuditLog).filter(AuditLog.org_id == org_id).all()
        
        return {
            "org_name": org_name,
            "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "suppliers": suppliers,
            "alerts": alerts,
            "plans": plans,
            "audits": audits,
            "report_type": report_type.upper()
        }

    # 1. PDF GENERATOR USING REPORTLAB
    @classmethod
    def generate_pdf(cls, db: Session, org_id: str, report_type: str) -> bytes:
        data = cls._get_report_data(db, org_id, report_type)
        buffer = io.BytesIO()
        
        # Document Setup
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=54,
            leftMargin=54,
            topMargin=54,
            bottomMargin=72
        )
        
        # Style Sheet
        styles = getSampleStyleSheet()
        
        # Colors definition
        primary_color = colors.HexColor("#0F172A") # slate 900
        secondary_color = colors.HexColor("#0D9488") # teal 600
        text_color = colors.HexColor("#334155") # slate 700
        light_bg = colors.HexColor("#F8FAFC") # slate 50
        accent_red = colors.HexColor("#DC2626") # red 600
        
        # Custom Paragraph Styles
        styles.add(ParagraphStyle('DocTitle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=26, leading=30, textColor=primary_color))
        styles.add(ParagraphStyle('DocSub', parent=styles['Normal'], fontName='Helvetica', fontSize=12, leading=16, textColor=secondary_color))
        styles.add(ParagraphStyle('SecHeader', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=16, leading=20, textColor=primary_color, spaceBefore=18, spaceAfter=8))
        styles.add(ParagraphStyle('SecBody', parent=styles['Normal'], fontName='Helvetica', fontSize=10, leading=14, textColor=text_color, spaceAfter=10))
        styles.add(ParagraphStyle('TableText', parent=styles['Normal'], fontName='Helvetica', fontSize=9, leading=11, textColor=text_color))
        styles.add(ParagraphStyle('TableHeader', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, leading=11, textColor=colors.white))
        
        story = []
        
        # --- TITLE BANNER ---
        story.append(Paragraph(f"SupplyVision AI Decisions Digest", styles['DocSub']))
        story.append(Paragraph(f"{report_type.replace('_', ' ').title()} Report", styles['DocTitle']))
        story.append(Paragraph(f"Client Organization: <b>{data['org_name']}</b> | Run Time: {data['date']}", styles['SecBody']))
        story.append(Spacer(1, 15))
        
        # --- HORIZONTAL DIVIDER ---
        divider = Table([[""]], colWidths=[504], rowHeights=[2])
        divider.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), secondary_color),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
        ]))
        story.append(divider)
        story.append(Spacer(1, 15))
        
        # --- CONDITIONAL CONTENT GENERATION ---
        if report_type == "executive":
            # Overview Metrics
            total_suppliers = len(data["suppliers"])
            total_alerts = len(data["alerts"])
            open_alerts = len([a for a in data["alerts"] if a.status == "open"])
            rupees_at_risk = sum(a.rupees_at_risk for a in data["alerts"] if a.status == "open")
            accepted_plans = len([p for p in data["plans"] if p.accepted_option_idx is not None])
            
            story.append(Paragraph("Resilience Metrics Overview", styles['SecHeader']))
            story.append(Paragraph("This brief highlights the overall resilience metrics of the supply chain including exposure and resolved contingencies.", styles['SecBody']))
            
            # KPI Card Grid Table
            kpi_data = [
                [
                    Paragraph("<b>Total Suppliers</b>", styles['TableText']),
                    Paragraph("<b>Threat Alerts</b>", styles['TableText']),
                    Paragraph("<b>Open Disruptions</b>", styles['TableText']),
                    Paragraph("<b>Rupees at Risk</b>", styles['TableText'])
                ],
                [
                    Paragraph(f"<font size=14><b>{total_suppliers}</b></font>", styles['TableText']),
                    Paragraph(f"<font size=14><b>{total_alerts}</b></font>", styles['TableText']),
                    Paragraph(f"<font size=14><b>{open_alerts}</b></font>", styles['TableText']),
                    Paragraph(f"<font size=14 color='#DC2626'><b>₹{rupees_at_risk:,.0f}</b></font>", styles['TableText'])
                ]
            ]
            kpi_table = Table(kpi_data, colWidths=[126, 126, 126, 126])
            kpi_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), light_bg),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#CBD5E1")),
                ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
                ('TOPPADDING', (0,0), (-1,-1), 10),
                ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ]))
            story.append(kpi_table)
            story.append(Spacer(1, 15))
            
            # Recent Ingested Signals Audit
            story.append(Paragraph("Logistics Security Alerts", styles['SecHeader']))
            story.append(Paragraph("Active alerts generated by external hazard feeds:", styles['SecBody']))
            
            alert_headers = [Paragraph("<b>Node ID</b>", styles['TableHeader']), 
                             Paragraph("<b>Type</b>", styles['TableHeader']), 
                             Paragraph("<b>Risk</b>", styles['TableHeader']), 
                             Paragraph("<b>Exposure (₹)</b>", styles['TableHeader']), 
                             Paragraph("<b>Status</b>", styles['TableHeader'])]
            alert_rows = [alert_headers]
            
            for a in data["alerts"][:6]:
                status_color = "#DC2626" if a.status == "open" else "#0D9488"
                alert_rows.append([
                    Paragraph(a.node_id, styles['TableText']),
                    Paragraph(a.node_type, styles['TableText']),
                    Paragraph(f"<b>{a.risk_score}</b>", styles['TableText']),
                    Paragraph(f"₹{a.rupees_at_risk:,.2f}", styles['TableText']),
                    Paragraph(f"<font color='{status_color}'><b>{a.status.replace('_', ' ').upper()}</b></font>", styles['TableText'])
                ])
                
            alert_table = Table(alert_rows, colWidths=[90, 80, 50, 164, 120])
            alert_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), primary_color),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_bg]),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
                ('TOPPADDING', (0,0), (-1,-1), 6),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ]))
            story.append(alert_table)
            
        elif report_type == "risk":
            story.append(Paragraph("Active Vulnerability Catalog", styles['SecHeader']))
            story.append(Paragraph("Details of all operational risk alerts flagged by OpenWeather, IMD, GDACS, and NewsAPI pipelines.", styles['SecBody']))
            
            rows = [[
                Paragraph("<b>Alert ID</b>", styles['TableHeader']),
                Paragraph("<b>Target Node</b>", styles['TableHeader']),
                Paragraph("<b>Type</b>", styles['TableHeader']),
                Paragraph("<b>Risk Score</b>", styles['TableHeader']),
                Paragraph("<b>Rupees at Risk</b>", styles['TableHeader'])
            ]]
            for a in data["alerts"]:
                rows.append([
                    Paragraph(a.id[:8] + "...", styles['TableText']),
                    Paragraph(a.node_id, styles['TableText']),
                    Paragraph(a.node_type, styles['TableText']),
                    Paragraph(f"<b>{a.risk_score}</b>", styles['TableText']),
                    Paragraph(f"₹{a.rupees_at_risk:,.2f}", styles['TableText'])
                ])
            t = Table(rows, colWidths=[80, 104, 80, 80, 160])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), primary_color),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_bg]),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ]))
            story.append(t)
            
        elif report_type == "recovery":
            story.append(Paragraph("Mitigation and Contingency Records", styles['SecHeader']))
            story.append(Paragraph("Breakdown of recommended alternate suppliers and accepted routing directives.", styles['SecBody']))
            
            rows = [[
                Paragraph("<b>Alert ID</b>", styles['TableHeader']),
                Paragraph("<b>Accepted Option</b>", styles['TableHeader']),
                Paragraph("<b>Approved By (User)</b>", styles['TableHeader']),
                Paragraph("<b>Timestamp</b>", styles['TableHeader'])
            ]]
            for p in data["plans"]:
                opt_title = "None"
                if p.accepted_option_idx is not None and isinstance(p.options_json, list):
                    opt_title = p.options_json[p.accepted_option_idx].get("title", "Alternative")
                rows.append([
                    Paragraph(p.alert_id[:8] + "...", styles['TableText']),
                    Paragraph(opt_title, styles['TableText']),
                    Paragraph(p.accepted_by or "N/A", styles['TableText']),
                    Paragraph(p.accepted_at.strftime("%Y-%m-%d %H:%M") if p.accepted_at else "Pending", styles['TableText'])
                ])
            t = Table(rows, colWidths=[100, 184, 110, 110])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), primary_color),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_bg]),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ]))
            story.append(t)
            
        elif report_type == "supplier":
            story.append(Paragraph("Digital Twin Nodes Ledger", styles['SecHeader']))
            story.append(Paragraph("Registered suppliers catalog with calculated reliability and risk ratings.", styles['SecBody']))
            
            rows = [[
                Paragraph("<b>Code</b>", styles['TableHeader']),
                Paragraph("<b>City / Hub</b>", styles['TableHeader']),
                Paragraph("<b>Tier</b>", styles['TableHeader']),
                Paragraph("<b>Reliability (%)</b>", styles['TableHeader']),
                Paragraph("<b>Exposure (₹)</b>", styles['TableHeader'])
            ]]
            for s in data["suppliers"]:
                rows.append([
                    Paragraph(s.get("id"), styles['TableText']),
                    Paragraph(s.get("city", "N/A"), styles['TableText']),
                    Paragraph(str(s.get("tier", 1)), styles['TableText']),
                    Paragraph(f"{s.get('reliability_score', 95)}%", styles['TableText']),
                    Paragraph(f"₹{s.get('revenue_exposure_inr', 0):,.2f}", styles['TableText'])
                ])
            t = Table(rows, colWidths=[100, 110, 60, 100, 134])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), primary_color),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_bg]),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ]))
            story.append(t)
            
        elif report_type == "monthly":
            story.append(Paragraph("Monthly Summary & Audits", styles['SecHeader']))
            story.append(Paragraph("Security and operational logs tracking user transactions over the past 30 days.", styles['SecBody']))
            
            rows = [[
                Paragraph("<b>Timestamp</b>", styles['TableHeader']),
                Paragraph("<b>Action Type</b>", styles['TableHeader']),
                Paragraph("<b>User ID</b>", styles['TableHeader']),
                Paragraph("<b>Details</b>", styles['TableHeader'])
            ]]
            for log in data["audits"][:20]:
                rows.append([
                    Paragraph(log.created_at.strftime("%Y-%m-%d %H:%M"), styles['TableText']),
                    Paragraph(log.action.upper(), styles['TableText']),
                    Paragraph(log.user_id[:8] if log.user_id else "SYSTEM", styles['TableText']),
                    Paragraph(log.resource_type or "General", styles['TableText'])
                ])
            t = Table(rows, colWidths=[110, 150, 94, 150])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), primary_color),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_bg]),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ]))
            story.append(t)

        # Build PDF
        doc.build(story, canvasmaker=NumberedCanvas)
        return buffer.getvalue()


    # 2. EXCEL GENERATOR USING OPENPYXL
    @classmethod
    def generate_excel(cls, db: Session, org_id: str, report_type: str) -> bytes:
        data = cls._get_report_data(db, org_id, report_type)
        wb = Workbook()
        ws = wb.active
        ws.title = report_type.replace('_', ' ').title()
        
        # Styles
        title_font = Font(name='Segoe UI', size=16, bold=True, color='0F172A')
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        data_font = Font(name='Segoe UI', size=11, color='334155')
        
        header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
        zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        # Title Block
        ws.append([f"SupplyVision AI - {report_type.replace('_', ' ').upper()} REPORT"])
        ws.cell(row=1, column=1).font = title_font
        ws.append([f"Org: {data['org_name']} | Generated: {data['date']}"])
        ws.append([]) # empty row
        
        if report_type == "executive":
            headers = ["Total Suppliers", "Threat Alerts", "Open Alerts", "Rupees at Risk"]
            ws.append(headers)
            
            total_suppliers = len(data["suppliers"])
            total_alerts = len(data["alerts"])
            open_alerts = len([a for a in data["alerts"] if a.status == "open"])
            rupees_at_risk = sum(a.rupees_at_risk for a in data["alerts"] if a.status == "open")
            
            ws.append([total_suppliers, total_alerts, open_alerts, rupees_at_risk])
            ws.cell(row=5, column=4).number_format = '₹#,##0.00'
            
        elif report_type == "risk":
            headers = ["Alert ID", "Node ID", "Node Type", "Risk Score", "Rupees at Risk", "Status", "Created At"]
            ws.append(headers)
            for a in data["alerts"]:
                ws.append([a.id, a.node_id, a.node_type, a.risk_score, a.rupees_at_risk, a.status, a.created_at.strftime("%Y-%m-%d %H:%M")])
                
        elif report_type == "recovery":
            headers = ["Plan ID", "Alert ID", "Org ID", "Accepted Option Index", "Accepted By (User)", "Accepted At", "Created At"]
            ws.append(headers)
            for p in data["plans"]:
                ws.append([p.id, p.alert_id, p.org_id, p.accepted_option_idx, p.accepted_by, p.accepted_at, p.created_at])
                
        elif report_type == "supplier":
            headers = ["Supplier ID", "City", "State", "Tier", "Reliability Score (%)", "Revenue Exposure (INR)"]
            ws.append(headers)
            for s in data["suppliers"]:
                ws.append([s.get("id"), s.get("city"), s.get("state"), s.get("tier"), s.get("reliability_score"), s.get("revenue_exposure_inr")])
                
        elif report_type == "monthly":
            headers = ["Log ID", "Org ID", "User ID", "Action", "Resource Type", "Resource ID", "Created At"]
            ws.append(headers)
            for log in data["audits"]:
                ws.append([log.id, log.org_id, log.user_id, log.action, log.resource_type, log.resource_id, log.created_at])
                
        # Format Table
        headers_row_idx = 4
        # Apply fonts and cell fills
        for row in range(headers_row_idx, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if row == headers_row_idx:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.font = data_font
                    cell.border = thin_border
                    if row % 2 == 0:
                        cell.fill = zebra_fill
                        
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()


    # 3. CSV GENERATOR
    @classmethod
    def generate_csv(cls, db: Session, org_id: str, report_type: str) -> str:
        data = cls._get_report_data(db, org_id, report_type)
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow([f"SupplyVision AI - {report_type.replace('_', ' ').upper()} REPORT"])
        writer.writerow([f"Org: {data['org_name']}", f"Generated: {data['date']}"])
        writer.writerow([])
        
        if report_type == "executive":
            writer.writerow(["Total Suppliers", "Threat Alerts", "Open Alerts", "Rupees at Risk"])
            total_suppliers = len(data["suppliers"])
            total_alerts = len(data["alerts"])
            open_alerts = len([a for a in data["alerts"] if a.status == "open"])
            rupees_at_risk = sum(a.rupees_at_risk for a in data["alerts"] if a.status == "open")
            writer.writerow([total_suppliers, total_alerts, open_alerts, rupees_at_risk])
            
        elif report_type == "risk":
            writer.writerow(["Alert ID", "Node ID", "Node Type", "Risk Score", "Rupees at Risk", "Status", "Created At"])
            for a in data["alerts"]:
                writer.writerow([a.id, a.node_id, a.node_type, a.risk_score, a.rupees_at_risk, a.status, a.created_at])
                
        elif report_type == "recovery":
            writer.writerow(["Plan ID", "Alert ID", "Org ID", "Accepted Option Index", "Accepted By (User)", "Accepted At", "Created At"])
            for p in data["plans"]:
                writer.writerow([p.id, p.alert_id, p.org_id, p.accepted_option_idx, p.accepted_by, p.accepted_at, p.created_at])
                
        elif report_type == "supplier":
            writer.writerow(["Supplier ID", "City", "State", "Tier", "Reliability Score (%)", "Revenue Exposure (INR)"])
            for s in data["suppliers"]:
                writer.writerow([s.get("id"), s.get("city"), s.get("state"), s.get("tier"), s.get("reliability_score"), s.get("revenue_exposure_inr")])
                
        elif report_type == "monthly":
            writer.writerow(["Log ID", "Org ID", "User ID", "Action", "Resource Type", "Resource ID", "Created At"])
            for log in data["audits"]:
                writer.writerow([log.id, log.org_id, log.user_id, log.action, log.resource_type, log.resource_id, log.created_at])
                
        return output.getvalue()
